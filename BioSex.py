import tkinter as tk
from tkinter import font
from tkinter import ttk
from PIL import Image, ImageTk
import csv
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import whisper
import pyaudio #CAPTURA EN TIEMPO REAL EL AUDIO 
import wave #PARA IMPORTAR EL AUDIO EN WAV

#--------------------------------------------PARTE DE GABRIEL--------------------------------------------

# Archivo Excel para guardar los datos
EXCEL_FILE = "Histo.xlsx"

# Variables globales para las enfermedades
clamidia = 0
gonorrea = 0
sifilis = 0
herpes = 0
sida = 0
vph = 0

variables_respuestas = []

def inicializar_excel():
    """Verifica o inicializa el archivo Excel."""
    try:
        workbook = load_workbook(EXCEL_FILE)
        if "Respuestas" not in workbook.sheetnames:
            hoja = workbook.create_sheet("Respuestas")
            _crear_encabezados(hoja)
    except FileNotFoundError:
        workbook = Workbook()
        hoja = workbook.active
        hoja.title = "Respuestas"
        _crear_encabezados(hoja)
    workbook.save(EXCEL_FILE)

def _crear_encabezados(hoja):
    """Crea los encabezados en la hoja."""
    encabezados = [
        "Nombre", "Apellido", "Edad", "Género", 
        "Pregunta 1", "Pregunta 2", "Pregunta 3", 
        "Pregunta 4", "Pregunta 5", "Pregunta 6", 
        "Pregunta 7", "Pregunta 8", "Pregunta 9", 
        "Pregunta 10", "Pregunta 11", "Pregunta 12"
    ]
    for col, encabezado in enumerate(encabezados, 1):
        hoja.cell(row=1, column=col, value=encabezado)
    
    # Ajustar tamaños de columna
    hoja.column_dimensions['A'].width = 20  # Nombre
    hoja.column_dimensions['B'].width = 20  # Apellido
    hoja.column_dimensions['C'].width = 10  # Edad
    hoja.column_dimensions['D'].width = 12  # Género
    for col in range(5, 17):  # Preguntas 1 a 12
        hoja.column_dimensions[get_column_letter(col)].width = 8  # Ancho reducido

# Llamar a la función para inicializar el archivo Excel
inicializar_excel()

#--------------------------------------------PARTE DE MELISSA--------------------------------------------

def escuchar_y_procesar_respuesta():
    modelo = whisper.load_model("base") #más ligera y rápida 
    CHUNK = 1024 #tamaño de los bloques de audio que se capturan por fragmentos 
    FORMAT = pyaudio.paInt16 #el audio se guarda en 16 bts 
    CHANNELS = 1 #graba en modo monofonico
    RATE = 16000 #frecuencia
    DURACION = 5
    ARCHIVO_SALIDA = "respuesta.wav"

    # Configurar PyAudio
    audio = pyaudio.PyAudio()
    stream = audio.open(format=FORMAT, channels=CHANNELS, rate=RATE, input=True, frames_per_buffer=CHUNK)
    
    print("Escuchando...")
    frames = []
    for _ in range(0, int(RATE / CHUNK * DURACION)):
        data = stream.read(CHUNK)
        frames.append(data)
    
    # Detener y guardar la grabación
    stream.stop_stream()
    stream.close()
    audio.terminate()

    with wave.open(ARCHIVO_SALIDA, 'wb') as wf:
        wf.setnchannels(CHANNELS)
        wf.setsampwidth(audio.get_sample_size(FORMAT))
        wf.setframerate(RATE)
        wf.writeframes(b''.join(frames))
    
    print("Procesando respuesta...")
    resultado = modelo.transcribe(ARCHIVO_SALIDA, language="es")  # Whisper transcribe
    texto = resultado['text'].strip().lower()
    
    if "sí" or "Si" or "sii" or "siii" or "siiii" or "siiii" or "siiiii" or "seee" or "see" or "seeee" or "se" in texto:
        return 1  # Equivalente a "Sí"
    elif "no" or "noo" or "nooo" or "noooo" or "nooooo" or "noooooo" in texto:
        return 0  # Equivalente a "No"
    else:
        print("No se entendió, intente nuevamente.")
        return None  # No entendió

#--------------------------------------------PARTE DE GABRIEL--------------------------------------------

def guardar_en_excel(nombre, apellido, edad, genero, respuestas):
    """Guarda los datos en el archivo Excel."""
    workbook = load_workbook(EXCEL_FILE)
    hoja = workbook["Respuestas"]
    nueva_fila = [nombre, apellido, edad, genero] + respuestas
    hoja.append(nueva_fila)
    workbook.save(EXCEL_FILE)


def mostrar_historial():
    """Muestra los datos guardados en el archivo Excel en una tabla más compacta."""
    try:
        workbook = load_workbook(EXCEL_FILE)
        hoja = workbook["Respuestas"]
        ventana_historial = tk.Toplevel()
        ventana_historial.title("Historial")
        ventana_historial.attributes('-fullscreen', True)

        # Crear un widget Treeview para mostrar los datos
        tree = ttk.Treeview(ventana_historial, columns=("Nombre", "Apellido", "Edad", "Género", "P1", "P2", "P3", "P4", "P5", "P6", "P7", "P8", "P9", "P10", "P11", "P12"), show="headings")
        
        # Definir el tamaño de las columnas
        tree.column("Nombre", width=100, anchor="center")
        tree.column("Apellido", width=100, anchor="center")
        tree.column("Edad", width=50, anchor="center")
        tree.column("Género", width=80, anchor="center")
        for i in range(1, 13):  # Columnas de Pregunta 1 a Pregunta 12
            tree.column(f"P{i}", width=80, anchor="center")
        
        # Agregar los encabezados
        tree.heading("Nombre", text="Nombre")
        tree.heading("Apellido", text="Apellido")
        tree.heading("Edad", text="Edad")
        tree.heading("Género", text="Género")
        for i in range(1, 13):
            tree.heading(f"P{i}", text=f"Pregunta {i}")
        
        # Insertar los datos de Excel en la tabla
        for row in hoja.iter_rows(values_only=True):
            tree.insert("", "end", values=row)

        # Hacer que el Treeview ocupe todo el espacio disponible
        tree.pack(expand=True, fill="both", padx=10, pady=10)

        # Botón para cerrar la ventana y volver a la ventana principal
        boton_volver = tk.Button(ventana_historial, text="Volver", command=ventana_historial.destroy, font=("Verdana", 14, "bold"), bg="#3E0283", fg="white")
        boton_volver.place(relx=0.5, rely=0.9, anchor="center", width=150, height=50)

    except FileNotFoundError:
        print("No se encontró el archivo de historial.")
def actualizar_enfermedad(enfermedad):
    """Suma 1 a la enfermedad correspondiente."""
    global clamidia, gonorrea, sifilis, herpes, sida, vph
    if enfermedad == "clamidia":
        clamidia += 1
    elif enfermedad == "gonorrea":
        gonorrea += 1
    elif enfermedad == "sifilis":
        sifilis += 1
    elif enfermedad == "herpes":
        herpes += 1
    elif enfermedad == "sida":
        sida += 1
    elif enfermedad == "vph":
        vph += 1

#--------------------------------------------PARTE DE MARIA DEL ROSARIO--------------------------------------------

def que1():
    ventana_pregunta(
        imagen_fondo="Q1.jpg",  # Imagen específica para esta pregunta
        enfermedad="clamidia",  # Variable a modificar
        siguiente_funcion=que2,  # Función para la siguiente ventana
        imagen_info="info1.jpg",  # Imagen para la subventana de información
        indice_pregunta=0
    )

def que2():
    ventana_pregunta(
        imagen_fondo="Q2.jpg",
        enfermedad="clamidia",
        siguiente_funcion=que3,
        imagen_info="info2.jpg",
        indice_pregunta=1
    )

def que3():
    ventana_pregunta(
        imagen_fondo="Q3.jpg",
        enfermedad="gonorrea",
        siguiente_funcion=que4,
        imagen_info="info3.jpg",
        indice_pregunta=2
    )

def que4():
    ventana_pregunta(
        imagen_fondo="Q4.jpg",
        enfermedad="sifilis",
        siguiente_funcion=que5,
        imagen_info="info4.jpg",
        indice_pregunta=3
    )

def que5():
    ventana_pregunta(
        imagen_fondo="Q5.jpg",
        enfermedad="sida",
        siguiente_funcion=que6,
        imagen_info="info5.jpg",
        indice_pregunta=4
    )

def que6():
    ventana_pregunta(
        imagen_fondo="Q6.jpg",
        enfermedad="vph",
        siguiente_funcion=que7,
        imagen_info="info6.jpg",
        indice_pregunta=5
    )

def que7():
    ventana_pregunta(
        imagen_fondo="Q7.jpg",
        enfermedad="sida",
        siguiente_funcion=que8,
        imagen_info="info7.jpg",
        indice_pregunta=6
    )

def que8():
    ventana_pregunta(
        imagen_fondo="Q8.jpg",
        enfermedad="gonorrea",
        siguiente_funcion=que9,
        imagen_info="info8.jpg",
        indice_pregunta=7
    )

def que9():
    ventana_pregunta(
        imagen_fondo="Q9.jpg",
        enfermedad="herpes",
        siguiente_funcion=que10,
        imagen_info="info9.jpg",
        indice_pregunta=8
    )

def que10():
    ventana_pregunta(
        imagen_fondo="Q10.jpg",
        enfermedad="sifilis",
        siguiente_funcion=que11,
        imagen_info="info10.jpg",
        indice_pregunta=9
    )

def que11():
    ventana_pregunta(
        imagen_fondo="Q11.jpg",
        enfermedad="herpes",
        siguiente_funcion=que12,
        imagen_info="info11.jpg",
        indice_pregunta=10
    )

def que12():
    ventana_pregunta(
        imagen_fondo="Q12.jpg",
        enfermedad="vph",
        siguiente_funcion=mostrar_resultados,  # Última ventana, muestra resultados
        imagen_info="info12.jpg",
        indice_pregunta=11
    )

#--------------------------------------------PARTE DE MELISSA--------------------------------------------

def mostrar_resultados():
    ventana_resultados = tk.Toplevel()
    ventana_resultados.attributes('-fullscreen', True)

    # Intentar cargar la imagen de fondo
    try:
        imagen_fondo = Image.open("resu.jpg")  
        imagen_fondo = imagen_fondo.resize(
            (ventana_resultados.winfo_screenwidth(), ventana_resultados.winfo_screenheight()),
            Image.Resampling.LANCZOS
        )
        fondo = ImageTk.PhotoImage(imagen_fondo)
        label_fondo = tk.Label(ventana_resultados, image=fondo)
        label_fondo.place(relwidth=1, relheight=1)  # Fondo ocupa toda la ventana
        ventana_resultados.imagen_fondo = fondo 
    except Exception as e:
        print(f"No se pudo cargar la imagen de fondo: {e}")
        label_fondo = tk.Label(ventana_resultados, bg="lightblue")  # Fondo alternativo
        label_fondo.place(relwidth=1, relheight=1)

    # Mostrar los resultados (enfermedades mayores a un valor a 1)
    resultado_texto = "Posibles Enfermedades Detectadas:\n"
    enfermedades_detectadas = []

    if clamidia == 2:
        enfermedades_detectadas.append("•CLAMIDIA: Hacerse pruebas cada cierto tiempo y usar condones cuando tengas sexo.")
    if gonorrea == 2:
        enfermedades_detectadas.append("•GONORREA: Usar protección y acudir al médico si presentas síntomas.")
    if sifilis == 2:
        enfermedades_detectadas.append("•SIFILIS: Realizarse exámenes regulares y consultar al médico.")
    if herpes == 2:
        enfermedades_detectadas.append("•HERPES: Evitar contacto directo durante los brotes y usar protección.")
    if sida == 2:
        enfermedades_detectadas.append("•SIDA: Realizarse chequeos periódicos y consultar sobre tratamientos.")
    if vph == 2:
        enfermedades_detectadas.append("•VPH: Es importante realizarse pruebas y llevar un seguimiento médico.")

    # Si no hay resultados, mostrar un mensaje alternativo
    if not enfermedades_detectadas:
        resultado_texto = "¡No se detectaron enfermedades!"
        label_resultados = tk.Label(
            ventana_resultados, text=resultado_texto, font=("Verdana", 18, "bold"),
            justify="center", bg="#3E0283", fg="white"
        )
        label_resultados.place(relx=0.5, rely=0.4, anchor="center")
    else:
        # Título centrado
        label_titulo = tk.Label(
            ventana_resultados, text="POSIBLES ENFERMEDADES DETECTADAS", font=("Verdana", 24, "bold"),
            bg="#3E0283", fg="white", justify="center"
        )
        label_titulo.place(relx=0.5, rely=0.25, anchor="center")  # Más abajo

        # Mostrar las enfermedades con consejos debajo, alineadas a la izquierda
        y_offset = 0.35  # Punto de inicio vertical para el texto
        for enfermedad in enfermedades_detectadas:
            label_enfermedad = tk.Label(
                ventana_resultados, text=enfermedad, font=("Verdana", 16, "bold"),
                bg="#3E0283", fg="white", justify="left", anchor="w"
            )
            label_enfermedad.place(relx=0.2, rely=y_offset, anchor="w")  # Alinear a la izquierda
            y_offset += 0.08  # Incrementar la posición vertical para la siguiente línea

    # Función para cerrar la ventana de resultados y mostrar la ventana principal
    def regresar_a_principal():
        ventana_resultados.destroy()  # Cierra la ventana de resultados
        ventana_principal_instancia.deiconify()  # Muestra la ventana principal nuevamente

     # Botón para regresar al inicio
    boton_regresar = tk.Button(
        ventana_resultados,
        text="REGRESAR AL MENÚ",
        bg="#3E0283",
        fg="white",
        font=("Verdana", 16, "bold"),
        relief="flat",
        command=regresar_a_principal
    )
    boton_regresar.place(relx=0.49, rely=0.868, anchor="center", width=234, height=48)

#--------------------------------------------PARTE DE HORACIO--------------------------------------------

def ventana_pregunta(imagen_fondo, enfermedad, siguiente_funcion, imagen_info, indice_pregunta):
    """Función genérica para manejar ventanas de preguntas."""
    ventana = tk.Toplevel()
    ventana.attributes('-fullscreen', True)

    # Cargar la imagen de fondo
    try:
        imagen = Image.open(imagen_fondo)
        imagen = imagen.resize(
            (ventana.winfo_screenwidth(), ventana.winfo_screenheight()),
            Image.Resampling.LANCZOS
        )
        ventana.imagen_fondo = ImageTk.PhotoImage(imagen)
        label_fondo = tk.Label(ventana, image=ventana.imagen_fondo)
        label_fondo.place(relwidth=1, relheight=1)
    except Exception as e:
        print(f"No se pudo cargar la imagen {imagen_fondo}: {e}")
        label_fondo = tk.Label(ventana, bg="lightblue")
        label_fondo.place(relwidth=1, relheight=1)


    # Función para actualizar la respuesta y pasar a la siguiente ventana
    def responder(valor):
        # Actualiza la respuesta en variables_respuestas
        variables_respuestas[indice_pregunta].set(valor)
        print(f"Respuesta para la Pregunta {indice_pregunta + 1}: {'Sí' if valor == 1 else 'No'}")  # Solo depuración

        # Solo actualiza las enfermedades si la respuesta es "Sí" (valor = 1)
        if valor == 1:
            actualizar_enfermedad(enfermedad)

        # Cierra la ventana y pasa a la siguiente pregunta
        ventana.destroy()
        siguiente_funcion()

    # Botones de respuesta
    boton_si = tk.Button(
        ventana,
        text="SI",
        bg="#008000",
        fg="white",
        font=("Verdana", 16, "bold"),
        command=lambda: responder(1)  # 1 = Sí
    )
    boton_si.place(relx=0.4, rely=0.5, anchor="center")

    boton_no = tk.Button(
        ventana,
        text="NO",
        bg="#FF0000",
        fg="white",
        font=("Verdana", 16, "bold"),
        command=lambda: responder(0)  # 0 = No
    )
    boton_no.place(relx=0.6, rely=0.5, anchor="center")

    # Botón de información
    def abrir_subventana():
        subventana = tk.Toplevel(ventana)
        subventana.transient(ventana)
        subventana.grab_set()
        subventana.attributes('-fullscreen', False)
        subventana.geometry("600x400")
        subventana.title("Información")

        # Cargar la imagen de fondo para la subventana
        try:
            info_imagen = Image.open(imagen_info)
            info_imagen = info_imagen.resize((600, 400), Image.Resampling.LANCZOS)
            subventana.imagen_info = ImageTk.PhotoImage(info_imagen)
            label_info = tk.Label(subventana, image=subventana.imagen_info)
            label_info.place(relwidth=1, relheight=1)
        except Exception as e:
            print(f"No se pudo cargar la imagen de información: {e}")
            subventana.configure(bg="white")

        # Botón para cerrar la subventana
        boton_cerrar = tk.Button(
            subventana, 
            text="CERRAR", 
            bg="white", 
            fg="#3E0283", 
            font=("Verdana", 12, "bold"), 
            relief="flat",
            command=subventana.destroy
        )
        boton_cerrar.place(relx=0.5, rely=0.85, anchor="center")

    boton_info = tk.Button(
        ventana,
        text="¿INFORMACIÓN?",
        bg="#3E0283",
        fg="white",
        font=("Verdana", 16, "bold"),
        relief="flat",  # Sin sombra en el contorno
        command=abrir_subventana
    )
    boton_info.place(relx=0.5, rely=0.85, anchor="center")

    # Función para manejar la respuesta de voz
    def responder_por_voz():
        valor = escuchar_y_procesar_respuesta()
        if valor is not None:
            responder(valor)
        else:
            print("No se entendió la respuesta, inténtalo de nuevo.")

    # Botón de respuesta por voz
    boton_voz = tk.Button(
        ventana,
        text="Responder por Voz",
        bg="#3E0283",
        fg="white",
        font=("Verdana", 16, "bold"),
        command=responder_por_voz
    )
    boton_voz.place(relx=0.5, rely=0.7, anchor="center")

#--------------------------------------------PARTE DE GABRIEL--------------------------------------------

def ventana_toma_datos():
    global variables_respuestas
    variables_respuestas = [tk.IntVar(value=0) for _ in range(12)]  # 12 preguntas, todas inicializadas en 0
    ventana_datos = tk.Toplevel()
    ventana_datos.attributes('-fullscreen', True)

    # Cargar la imagen de fondo
    try:
        imagen_fondo = Image.open("Datos.jpg")  # Imagen de fondo para la ventana
        imagen_fondo = imagen_fondo.resize(
            (ventana_datos.winfo_screenwidth(), ventana_datos.winfo_screenheight()),
            Image.Resampling.LANCZOS
        )
        ventana_datos.imagen_fondo = ImageTk.PhotoImage(imagen_fondo)
        label_fondo = tk.Label(ventana_datos, image=ventana_datos.imagen_fondo)
        label_fondo.place(relwidth=1, relheight=1)
    except Exception as e:
        print(f"No se pudo cargar la imagen de fondo: {e}")
        label_fondo = tk.Label(ventana_datos, bg="lightblue")
        label_fondo.place(relwidth=1, relheight=1)

    # Crear una fuente personalizada para las entradas
    fuente_entradas = font.Font(family="Verdana", size=12, weight="bold")

    # Entradas de datos
    entrada_nombre = tk.Entry(ventana_datos, font=fuente_entradas)
    entrada_nombre.place(relx=0.5, rely=0.4, anchor="center", width=300, height=40)

    entrada_apellido = tk.Entry(ventana_datos, font=fuente_entradas)
    entrada_apellido.place(relx=0.5, rely=0.5, anchor="center", width=300, height=40)

    entrada_edad = tk.Entry(ventana_datos, font=fuente_entradas)
    entrada_edad.place(relx=0.5, rely=0.6, anchor="center", width=300, height=40)

    # Opción desplegable para el género
    opciones_genero = ["Varón ♂️", "Mujer ♀️"]
    seleccion_genero = tk.StringVar(value="Seleccionar género")
    entrada_genero = ttk.Combobox(
        ventana_datos, textvariable=seleccion_genero, values=opciones_genero, font=fuente_entradas, state="readonly"
    )
    entrada_genero.place(relx=0.5, rely=0.7, anchor="center", width=300, height=40)

    # Función para validar que la edad sea un número
    def es_numero(valor):
        """Verifica si el valor ingresado es un número entero."""
        return valor.isdigit()

    # Función para guardar los datos
    def guardar_datos():
        nombre = entrada_nombre.get()
        apellido = entrada_apellido.get()
        edad = entrada_edad.get()
        genero = entrada_genero.get()

        # Validar si la edad es numérica
        if not es_numero(edad):
            print("La edad debe ser un número válido.")
            mostrar_alerta("La edad debe ser un número válido.")  # Mostrar alerta
            return

        if nombre and apellido and genero != "Seleccionar género":
            respuestas = ["Sí" if var.get() == 1 else "No" for var in variables_respuestas]
            guardar_en_excel(nombre, apellido, edad, genero, respuestas)
            print("Datos guardados correctamente.")
            ventana_datos.destroy()
            que1()  # Abrir la primera pregunta
        else:
            print("Por favor, complete todos los campos.")
            mostrar_alerta()

    # Función para mostrar alerta si faltan datos
    def mostrar_alerta():
        ventana_alerta = tk.Toplevel(ventana_datos)
        ventana_alerta.title("Advertencia")
        ventana_alerta.geometry("400x200")
        ventana_alerta.resizable(False, False)
        ventana_alerta.configure(bg="red")
        ventana_alerta.transient(ventana_datos)  # Asociar ventana
        ventana_alerta.grab_set()  # Bloquear interacción
        ventana_alerta.focus_force()  # Enfocar ventana

        mensaje = tk.Label(
            ventana_alerta,
            text="¡Por favor complete todos los campos!",
            font=("Verdana", 14, "bold"),
            fg="white",
            bg="red",
            wraplength=350,
        )
        mensaje.pack(expand=True)

        boton_ok = tk.Button(
            ventana_alerta,
            text="OK",
            font=("Verdana", 12),
            command=ventana_alerta.destroy,
            bg="white",
            fg="black",
        )
        boton_ok.pack(pady=10)

    # Botón para guardar los datos
    boton_guardar = tk.Button(
        ventana_datos,
        text="Guardar Datos",
        font=fuente_entradas,
        command=guardar_datos
    )
    boton_guardar.place(relx=0.5, rely=0.8, anchor="center", width=200, height=50)

#--------------------------------------------PARTE DE HORACIO--------------------------------------------

def ventana_principal():
    # Crear la ventana principal
    ventana = tk.Tk()
    ventana.attributes('-fullscreen', True)  # Pantalla completa

    # Cargar la imagen de fondo
    try:
        imagen_fondo = Image.open("fondo.jpg")  
        imagen_fondo = imagen_fondo.resize(
            (ventana.winfo_screenwidth(), ventana.winfo_screenheight()),
            Image.Resampling.LANCZOS
        )
        fondo = ImageTk.PhotoImage(imagen_fondo)
        label_fondo = tk.Label(ventana, image=fondo)
        label_fondo.place(relwidth=1, relheight=1)  # Fondo ocupa toda la ventana
        ventana.imagen_fondo = fondo  # Guardamos la referencia de la imagen
    except Exception as e:
        print(f"No se pudo cargar la imagen de fondo: {e}")
        label_fondo = tk.Label(ventana, bg="lightblue")  # Fondo alternativo
        label_fondo.place(relwidth=1, relheight=1)

    # Fuente personalizada para los botones
    fuente_botones = font.Font(family="Verdana", size=18, weight="bold")

    # Marco para centrar los botones
    frame_botones = tk.Frame(ventana, bg="#3E0283")
    frame_botones.place(relx=0.5, rely=0.75, anchor="center")  # Posicionar en el centro inferior

    # Estilo de los botones
    estilo_botones = {
        "font": fuente_botones,
        "bg": "white",
        "fg": "black",
        "activebackground": "#A9A9A9",
        "activeforeground": "white",
        "relief": "raised",
        "padx": 20,
        "pady": 10,
    }

    # Botón de comenzar análisis
    btn_comenzar = tk.Button(
        frame_botones,
        text="¡COMENZAR ANÁLISIS!",
        **estilo_botones,
        command=lambda: [ventana.withdraw(), ventana_toma_datos()]
    )
    btn_comenzar.pack(pady=15)

    # Botón de historial
    btn_historial = tk.Button(
        frame_botones,
        text="VER HISTORIAL",
        **estilo_botones,
        command=mostrar_historial
    )
    btn_historial.pack(pady=15)

    # Botón de salir
    btn_salir = tk.Button(
        frame_botones,
        text="SALIR",
        **estilo_botones,
        command=ventana.destroy
    )
    btn_salir.pack(pady=15)

    # Retorna la ventana creada
    return ventana

# Llamamos a la función para crear la ventana principal
ventana_principal_instancia = ventana_principal()  # Obtenemos la instancia de la ventana

# Ejecutar el bucle principal de la ventana
ventana_principal_instancia.mainloop()  # Mantiene la ventana abierta hasta que el usuario la cierre
